"""
Utilidades para importar archivos Excel del PAC.
Formato del Excel: Estructura jerarquica con secciones INGRESOS y GASTOS.

Columnas:
  A: Numero RP/CxP (solo en reservas/cuentas por pagar)
  B: Codigo presupuestal
  C: Descripcion/nombre del rubro
  D: Apropiacion Inicial
  E: Adiciones
  F: Reduccion
  G: Creditos
  H: Contracreditos
  I: Apropiacion Definitiva
  J-U: Enero a Diciembre
  V: Total
"""

import re
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook


def es_item_hoja(codigo_str):
    """
    Determina si un codigo presupuestal es un item hoja (detalle) o un subtotal (padre).

    Los items hoja tienen 3+ segmentos separados por ' - ':
      '1003 - 2.1.1.01.01.001.01 - 20'  -> ['1003', '2.1.1...', '20'] = 3 partes = HOJA
      '1003 - 2.3.21.2102.1900.001.2.3.2.02.02.005 - 05' -> 3 partes = HOJA

    Los subtotales tienen solo 1-2 segmentos:
      '1003 - 2.1'          -> ['1003', '2.1'] = 2 partes = SUBTOTAL
      '1003 - 0101'         -> ['1003', '0101'] = 2 partes = SUBTOTAL
      '1003 - 2.3.21'       -> ['1003', '2.3.21'] = 2 partes = SUBTOTAL
      '1'                   -> ['1'] = 1 parte = SUBTOTAL
    """
    if not codigo_str:
        return False
    parts = codigo_str.strip().split(' - ')
    return len(parts) >= 3


def safe_decimal(value):
    """Convierte un valor a Decimal de forma segura."""
    if value is None or value == '' or value == '-':
        return Decimal('0')
    try:
        if isinstance(value, str):
            value = value.replace(',', '').replace('$', '').replace(' ', '')
        val = Decimal(str(value))
        return val
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


def detectar_seccion(codigo, nombre, fila_idx, seccion_actual):
    """
    Detecta en que seccion del presupuesto se encuentra una fila.
    Retorna: (tipo, categoria, es_subtotal, seccion_nueva)

    Seccion states: 'INGRESOS' -> 'GASTOS' -> 'RESERVAS' -> 'CXP'
    """
    nombre_upper = (nombre or '').strip().upper()
    codigo_str = str(codigo or '').strip()

    # ========== 1. DETECTAR TRANSICIONES DE SECCION ==========

    # Transicion a GASTOS
    if nombre_upper in ('GASTOS', 'GASTO'):
        return 'GASTO', '', True, 'GASTOS'

    # Transicion a RESERVAS PRESUPUESTALES
    if 'RESERVAS PRESUPUESTAL' in nombre_upper or 'RESERVA PRESUPUESTAL' in nombre_upper:
        return 'GASTO', 'RESERVAS', True, 'RESERVAS'

    # Transicion a CUENTAS POR PAGAR (header puede ser C="CUENTAS POR PAGAR" o B="5")
    if 'CUENTAS POR PAGAR' in nombre_upper or (codigo_str == '5' and 'CUENTAS' in nombre_upper):
        return 'GASTO', 'CUENTAS_POR_PAGAR', True, 'CXP'

    # ========== 2. TOTALES Y SALDOS (cualquier seccion) ==========

    if codigo_str in ('A', 'B') or 'TOTAL INGRESOS' in nombre_upper or 'TOTAL GASTOS' in nombre_upper:
        tipo = 'INGRESO' if codigo_str == 'A' or 'INGRESO' in nombre_upper else 'GASTO'
        return tipo, '', True, seccion_actual

    if 'SALDO DISPONIBLE' in nombre_upper:
        return 'GASTO', '', True, seccion_actual

    # ========== 3. SECCION RESERVAS ==========

    if seccion_actual == 'RESERVAS':
        tipo = 'GASTO'
        if 'FUNCIONAMIENTO' in nombre_upper and not codigo_str:
            return tipo, 'RESERVAS', True, 'RESERVAS'
        if 'INVERSION' in nombre_upper and not codigo_str:
            return tipo, 'RESERVAS', True, 'RESERVAS'
        return tipo, 'RESERVAS', False, 'RESERVAS'

    # ========== 4. SECCION CUENTAS POR PAGAR ==========

    if seccion_actual == 'CXP':
        tipo = 'GASTO'
        if 'FUNCIONAMIENTO' in nombre_upper and not codigo_str:
            return tipo, 'CUENTAS_POR_PAGAR', True, 'CXP'
        if 'INVERSION' in nombre_upper and not codigo_str:
            return tipo, 'CUENTAS_POR_PAGAR', True, 'CXP'
        return tipo, 'CUENTAS_POR_PAGAR', False, 'CXP'

    # ========== 5. SECCION INGRESOS ==========

    if seccion_actual == 'INGRESOS':
        tipo = 'INGRESO'

        # Saldo inicial (codigo 1)
        if codigo_str == '1' or 'SALDO INICIAL' in nombre_upper:
            return tipo, 'SALDO_INICIAL', True, seccion_actual
        if 'CAJA' in nombre_upper or 'BANCOS' in nombre_upper:
            if not codigo_str or codigo_str in ('1.1', '1.2', '1.3'):
                return tipo, 'SALDO_INICIAL', False, seccion_actual

        # Ingresos corrientes (codigo 2)
        if codigo_str == '2' or 'INGRESOS CORRIENTES' in nombre_upper:
            return tipo, 'INGRESO_CORRIENTE', True, seccion_actual
        if 'TRIBUTARIO' in nombre_upper or 'NO TRIBUTARIO' in nombre_upper:
            return tipo, 'INGRESO_CORRIENTE', True, seccion_actual

        # Ingresos de capital (codigo 3)
        if codigo_str == '3' or 'INGRESOS DE CAPITAL' in nombre_upper or 'INGRESOS CAPITAL' in nombre_upper:
            return tipo, 'INGRESO_CAPITAL', True, seccion_actual

        # Detalle de ingresos con codigo presupuestal
        if '1003' in codigo_str or (codigo_str and not codigo_str.isalpha()):
            cat = 'INGRESO_CORRIENTE'
            if any(x in nombre_upper for x in ['CAPITAL', 'SUPERAVIT', 'RENDIMIENTO']):
                cat = 'INGRESO_CAPITAL'
            return tipo, cat, False, seccion_actual

        return tipo, 'INGRESO_CORRIENTE', False, seccion_actual

    # ========== 6. SECCION GASTOS ==========

    tipo = 'GASTO'

    # Inversion (codigo 2.3) - check FIRST because 2.3.xx codes are most common
    # and can accidentally contain substrings like '2.1' or '2.2'
    if '- 2.3' in codigo_str or codigo_str.endswith('2.3'):
        es_sub = codigo_str.endswith('2.3') or 'INVERSION' in nombre_upper
        return tipo, 'INVERSION', es_sub, 'GASTOS'
    if '2.3' in codigo_str:
        return tipo, 'INVERSION', False, 'GASTOS'

    # Funcionamiento (codigo 2.1)
    if '- 2.1' in codigo_str or codigo_str.endswith('2.1'):
        es_sub = codigo_str.endswith('2.1') or 'FUNCIONAMIENTO' in nombre_upper
        return tipo, 'FUNCIONAMIENTO', es_sub, 'GASTOS'
    if '2.1' in codigo_str:
        return tipo, 'FUNCIONAMIENTO', False, 'GASTOS'

    # Servicio a la deuda (codigo 2.2) - only match explicit pattern or name
    if '- 2.2' in codigo_str:
        return tipo, 'DEUDA', False, 'GASTOS'
    if 'DEUDA' in nombre_upper:
        return tipo, 'DEUDA', True, 'GASTOS'
    if any(x in nombre_upper for x in ['AMORTIZACION', 'INTERESES Y OTROS']):
        return tipo, 'DEUDA', False, 'GASTOS'

    # Sectores de inversion (SECTOR MINAS, SECTOR EDUCACION, etc.)
    if 'SECTOR' in nombre_upper:
        return tipo, 'INVERSION', True, 'GASTOS'

    # BPIN de inversion
    if 'BPIN' in nombre_upper or 'BPIN' in codigo_str:
        return tipo, 'INVERSION', False, 'GASTOS'

    # Funcionamiento si es solo label sin codigo
    if 'FUNCIONAMIENTO' in nombre_upper and not codigo_str:
        return tipo, 'FUNCIONAMIENTO', True, 'GASTOS'
    if 'INVERSION' in nombre_upper and not codigo_str:
        return tipo, 'INVERSION', True, 'GASTOS'

    # Default para gastos
    return tipo, 'FUNCIONAMIENTO', False, seccion_actual


def importar_excel_pac(archivo, vigencia, modelo_class, usuario, nombre_hoja=None):
    """
    Importa un archivo Excel con formato PAC real.
    Lee las filas del Excel y crea registros en el modelo especificado.

    Args:
        archivo: archivo Excel subido
        vigencia: año de vigencia
        modelo_class: clase del modelo (AIMInicial, PACProgramado, etc.)
        usuario: usuario que realiza la carga
        nombre_hoja: nombre de la hoja a leer (None = primera hoja)

    Returns:
        count: numero de registros importados
    """
    wb = load_workbook(archivo, data_only=True)

    if nombre_hoja:
        # Buscar hoja por nombre parcial
        ws = None
        for sname in wb.sheetnames:
            if nombre_hoja.upper() in sname.upper():
                ws = wb[sname]
                break
        if ws is None:
            ws = wb.active
    else:
        ws = wb.active

    # Eliminar datos anteriores de la misma vigencia
    modelo_class.objects.filter(vigencia=vigencia).delete()

    count = 0
    seccion_actual = 'INGRESOS'  # Empezamos en seccion de ingresos

    # Iterar desde fila 5 (despues de titulos y encabezados)
    for row_idx in range(5, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value  # Numero RP/CxP
        col_b = ws.cell(row=row_idx, column=2).value  # Codigo
        col_c = ws.cell(row=row_idx, column=3).value  # Descripcion

        # Saltar filas completamente vacias
        if col_b is None and col_c is None:
            continue

        codigo = str(col_b or '').strip()
        nombre = str(col_c or '').strip()

        # Si no tiene ni codigo ni nombre significativo, saltar
        if not codigo and not nombre:
            continue

        # Saltar filas de firma/footer
        if any(x in nombre.upper() for x in ['SUBGERENTE', 'GERENTE', 'FIRMA', 'ELABOR']):
            continue

        # Detectar tipo y categoria
        tipo, categoria, es_subtotal, seccion_actual = detectar_seccion(
            codigo, nombre, row_idx, seccion_actual
        )

        # Refinar es_subtotal usando patron de fuente en el codigo.
        # Items hoja tienen sufijo de fuente (ej: "- 20", "- 03")
        # Items sin fuente con codigo son subtotales/padres.
        if codigo:
            if es_item_hoja(codigo):
                es_subtotal = False
            elif codigo not in ('A', 'B', '1', '2', '3', '4', '5'):
                es_subtotal = True

        # Leer valores numericos
        aprop_inicial = safe_decimal(ws.cell(row=row_idx, column=4).value)
        adiciones_val = safe_decimal(ws.cell(row=row_idx, column=5).value)
        reduccion_val = safe_decimal(ws.cell(row=row_idx, column=6).value)
        creditos_val = safe_decimal(ws.cell(row=row_idx, column=7).value)
        contracred_val = safe_decimal(ws.cell(row=row_idx, column=8).value)
        aprop_def = safe_decimal(ws.cell(row=row_idx, column=9).value)

        enero_val = safe_decimal(ws.cell(row=row_idx, column=10).value)
        febrero_val = safe_decimal(ws.cell(row=row_idx, column=11).value)
        marzo_val = safe_decimal(ws.cell(row=row_idx, column=12).value)
        abril_val = safe_decimal(ws.cell(row=row_idx, column=13).value)
        mayo_val = safe_decimal(ws.cell(row=row_idx, column=14).value)
        junio_val = safe_decimal(ws.cell(row=row_idx, column=15).value)
        julio_val = safe_decimal(ws.cell(row=row_idx, column=16).value)
        agosto_val = safe_decimal(ws.cell(row=row_idx, column=17).value)
        septiembre_val = safe_decimal(ws.cell(row=row_idx, column=18).value)
        octubre_val = safe_decimal(ws.cell(row=row_idx, column=19).value)
        noviembre_val = safe_decimal(ws.cell(row=row_idx, column=20).value)
        diciembre_val = safe_decimal(ws.cell(row=row_idx, column=21).value)
        total_val = safe_decimal(ws.cell(row=row_idx, column=22).value)

        # Solo saltar filas sin datos numericos Y sin codigo significativo
        tiene_datos = any([
            aprop_inicial, adiciones_val, reduccion_val, creditos_val,
            contracred_val, aprop_def, enero_val, febrero_val, marzo_val,
            abril_val, mayo_val, junio_val, julio_val, agosto_val,
            septiembre_val, octubre_val, noviembre_val, diciembre_val, total_val
        ])

        if not tiene_datos and not codigo:
            continue

        # Si es solo el titulo de seccion "GASTOS", no crear registro
        if nombre.upper() == 'GASTOS' and not codigo:
            continue

        # Incluir numero RP/CxP en el codigo si existe
        if col_a and str(col_a).strip():
            rp_prefix = str(col_a).strip()
            if codigo:
                codigo = f"{codigo} (RP:{rp_prefix})"
            else:
                codigo = f"RP:{rp_prefix}"

        # Extraer fuente del codigo (ultimo segmento despues de " - ")
        fuente = ''
        if ' - ' in codigo:
            parts = codigo.split(' - ')
            if len(parts) >= 3:
                fuente_code = parts[-1].strip().split(' ')[0]  # Tomar solo el codigo
                fuente = fuente_code

        modelo_class.objects.create(
            vigencia=vigencia,
            tipo=tipo,
            categoria=categoria,
            codigo_rubro=codigo,
            nombre_rubro=nombre,
            fuente_financiacion=fuente,
            apropiacion_inicial=aprop_inicial,
            adiciones=adiciones_val,
            reduccion=reduccion_val,
            creditos=creditos_val,
            contracreditos=contracred_val,
            apropiacion_definitiva=aprop_def,
            enero=enero_val,
            febrero=febrero_val,
            marzo=marzo_val,
            abril=abril_val,
            mayo=mayo_val,
            junio=junio_val,
            julio=julio_val,
            agosto=agosto_val,
            septiembre=septiembre_val,
            octubre=octubre_val,
            noviembre=noviembre_val,
            diciembre=diciembre_val,
            total=total_val if total_val else (
                enero_val + febrero_val + marzo_val + abril_val +
                mayo_val + junio_val + julio_val + agosto_val +
                septiembre_val + octubre_val + noviembre_val + diciembre_val
            ),
            es_subtotal=es_subtotal,
            fila_excel=row_idx,
            usuario=usuario,
        )
        count += 1

    return count
