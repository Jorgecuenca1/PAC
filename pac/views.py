import json
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum
from django.db.models.functions import Coalesce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    AIMInicial, PACProgramado, PACEjecutadoCompromiso, PACEjecutadoPago,
    CargaArchivo, FuenteFinanciacion, MESES, MESES_DISPLAY
)
from .forms import ImportarArchivoForm, FuenteFinanciacionForm
from .utils import importar_excel_pac, safe_decimal


D0 = Decimal('0')


def _sum(qs, field):
    return qs.aggregate(t=Coalesce(Sum(field), D0))['t']


# ============================================================
# DASHBOARD
# ============================================================
@login_required
def dashboard(request):
    vigencia = int(request.GET.get('vigencia', 2026))

    # Excluir Saldo Inicial de ingresos y Reservas/CxP de gastos (vigencias anteriores)
    _excl_ing = ['SALDO_INICIAL']
    _excl_gas = ['RESERVAS', 'CUENTAS_POR_PAGAR']

    aim_qs = AIMInicial.objects.filter(vigencia=vigencia, es_subtotal=False)
    prog_qs = PACProgramado.objects.filter(vigencia=vigencia, es_subtotal=False)
    comp_qs = PACEjecutadoCompromiso.objects.filter(vigencia=vigencia, es_subtotal=False)
    pago_qs = PACEjecutadoPago.objects.filter(vigencia=vigencia, es_subtotal=False)

    aim_ingresos = _sum(aim_qs.filter(tipo='INGRESO').exclude(categoria__in=_excl_ing), 'apropiacion_definitiva')
    aim_gastos = _sum(aim_qs.filter(tipo='GASTO').exclude(categoria__in=_excl_gas), 'apropiacion_definitiva')

    prog_ingresos = _sum(prog_qs.filter(tipo='INGRESO').exclude(categoria__in=_excl_ing), 'total')
    prog_gastos = _sum(prog_qs.filter(tipo='GASTO').exclude(categoria__in=_excl_gas), 'total')

    comp_ingresos = _sum(comp_qs.filter(tipo='INGRESO').exclude(categoria__in=_excl_ing), 'total')
    comp_gastos = _sum(comp_qs.filter(tipo='GASTO').exclude(categoria__in=_excl_gas), 'total')

    pago_ingresos = _sum(pago_qs.filter(tipo='INGRESO').exclude(categoria__in=_excl_ing), 'total')
    pago_gastos = _sum(pago_qs.filter(tipo='GASTO').exclude(categoria__in=_excl_gas), 'total')

    datos_mensuales_ingresos = []
    datos_mensuales_gastos = []
    for mes in MESES:
        datos_mensuales_ingresos.append({
            'programado': float(_sum(prog_qs.filter(tipo='INGRESO').exclude(categoria__in=_excl_ing), mes)),
            'ejecutado': float(_sum(pago_qs.filter(tipo='INGRESO').exclude(categoria__in=_excl_ing), mes)),
        })
        datos_mensuales_gastos.append({
            'programado': float(_sum(prog_qs.filter(tipo='GASTO').exclude(categoria__in=_excl_gas), mes)),
            'ejecutado': float(_sum(pago_qs.filter(tipo='GASTO').exclude(categoria__in=_excl_gas), mes)),
        })

    pct_ing = (float(pago_ingresos) / float(prog_ingresos) * 100) if prog_ingresos else 0
    pct_gas = (float(pago_gastos) / float(prog_gastos) * 100) if prog_gastos else 0
    pct_comp = (float(pago_gastos) / float(comp_gastos) * 100) if comp_gastos else 0

    cargas_recientes = CargaArchivo.objects.all()[:5]

    context = {
        'vigencia': vigencia,
        'aim_ingresos': aim_ingresos, 'aim_gastos': aim_gastos,
        'prog_ingresos': prog_ingresos, 'prog_gastos': prog_gastos,
        'comp_ingresos': comp_ingresos, 'comp_gastos': comp_gastos,
        'pago_ingresos': pago_ingresos, 'pago_gastos': pago_gastos,
        'datos_mensuales_ingresos': json.dumps(datos_mensuales_ingresos),
        'datos_mensuales_gastos': json.dumps(datos_mensuales_gastos),
        'meses_display': json.dumps(MESES_DISPLAY),
        'pct_ing': round(pct_ing, 1), 'pct_gas': round(pct_gas, 1), 'pct_comp': round(pct_comp, 1),
        'cargas_recientes': cargas_recientes,
    }
    return render(request, 'pac/dashboard.html', context)


# ============================================================
# AIM INICIAL
# ============================================================
@login_required
def aim_inicial(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    tipo_filtro = request.GET.get('tipo', '')
    cat_filtro = request.GET.get('categoria', '')

    # AIM Inicial solo muestra presupuesto vigente (sin Saldo Inicial, Reservas, CxP)
    _excl_aim = ['SALDO_INICIAL', 'RESERVAS', 'CUENTAS_POR_PAGAR']
    registros = AIMInicial.objects.filter(vigencia=vigencia).exclude(categoria__in=_excl_aim)
    if tipo_filtro:
        registros = registros.filter(tipo=tipo_filtro)
    if cat_filtro:
        registros = registros.filter(categoria=cat_filtro)

    no_sub = registros.filter(es_subtotal=False)
    total_ingresos = _sum(no_sub.filter(tipo='INGRESO'), 'apropiacion_definitiva')
    total_gastos = _sum(no_sub.filter(tipo='GASTO'), 'apropiacion_definitiva')

    # Totales mensuales para el footer
    no_sub_all = registros.filter(es_subtotal=False)
    totales = {}
    for mes in MESES:
        totales[mes] = _sum(no_sub_all, mes)
    totales['total'] = _sum(no_sub_all, 'total')

    context = {
        'registros': registros,
        'vigencia': vigencia,
        'tipo_filtro': tipo_filtro,
        'cat_filtro': cat_filtro,
        'total_ingresos': total_ingresos,
        'total_gastos': total_gastos,
        'meses': MESES,
        'meses_display': MESES_DISPLAY,
        'totales': totales,
        'modulo': 'AIM Inicial',
        'color': '#ff9800',
    }
    return render(request, 'pac/aim_inicial.html', context)


@login_required
def importar_aim_inicial(request):
    if request.method == 'POST':
        form = ImportarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = form.cleaned_data['vigencia']
            try:
                count = importar_excel_pac(archivo, vigencia, AIMInicial, request.user)
                CargaArchivo.objects.create(
                    tipo='AIM_INICIAL', archivo=archivo, usuario=request.user,
                    registros_cargados=count,
                    observaciones=f'Vigencia {vigencia}. {count} registros cargados desde formato AIM.'
                )
                messages.success(request, f'Se cargaron {count} registros de AIM Inicial correctamente.')
                return redirect('aim_inicial')
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
    else:
        form = ImportarArchivoForm()
    context = {
        'form': form, 'modulo': 'AIM Inicial', 'color': '#ff9800',
        'tipo_importacion': 'AIM_INICIAL',
        'descripcion_formato': 'Archivo Excel con formato PAC de la entidad. Columnas: B=Codigo, C=Descripcion, D=Aprop.Inicial, E=Adiciones, F=Reduccion, G=Creditos, H=Contracred, I=Aprop.Definitiva, J-U=Meses, V=Total. Datos desde fila 5.',
    }
    return render(request, 'pac/importar.html', context)


# ============================================================
# PAC PROGRAMADO
# ============================================================
@login_required
def pac_programado(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    tipo_filtro = request.GET.get('tipo', '')
    cat_filtro = request.GET.get('categoria', '')

    registros = PACProgramado.objects.filter(vigencia=vigencia)
    if tipo_filtro:
        registros = registros.filter(tipo=tipo_filtro)
    if cat_filtro:
        registros = registros.filter(categoria=cat_filtro)

    no_sub = registros.filter(es_subtotal=False)
    totales = {}
    for mes in MESES:
        totales[mes] = _sum(no_sub, mes)
    totales['total'] = _sum(no_sub, 'total')

    context = {
        'registros': registros, 'vigencia': vigencia,
        'tipo_filtro': tipo_filtro, 'cat_filtro': cat_filtro,
        'totales': totales, 'meses': MESES, 'meses_display': MESES_DISPLAY,
        'modulo': 'PAC Programado', 'color': '#ffc107',
    }
    return render(request, 'pac/pac_mensual.html', context)


@login_required
def importar_pac_programado(request):
    if request.method == 'POST':
        form = ImportarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = form.cleaned_data['vigencia']
            try:
                count = importar_excel_pac(
                    archivo, vigencia, PACProgramado, request.user,
                    nombre_hoja='PROG PAC'
                )
                CargaArchivo.objects.create(
                    tipo='PROGRAMADO', archivo=archivo, usuario=request.user,
                    registros_cargados=count,
                    observaciones=f'Vigencia {vigencia}. {count} registros. Hoja: PROG PAC INGRESOS-GASTOS.'
                )
                messages.success(request, f'Se cargaron {count} registros de PAC Programado.')
                return redirect('pac_programado')
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
    else:
        form = ImportarArchivoForm()
    context = {
        'form': form, 'modulo': 'PAC Programado', 'color': '#ffc107',
        'tipo_importacion': 'PROGRAMADO',
        'descripcion_formato': 'Se lee la hoja "PROG PAC INGRESOS-GASTOS". Mismo formato del Excel PAC de la entidad.',
    }
    return render(request, 'pac/importar.html', context)


# ============================================================
# PAC EJECUTADO COMPROMISOS
# ============================================================
@login_required
def pac_ejecutado_compromisos(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    tipo_filtro = request.GET.get('tipo', '')
    cat_filtro = request.GET.get('categoria', '')

    registros = PACEjecutadoCompromiso.objects.filter(vigencia=vigencia)
    if tipo_filtro:
        registros = registros.filter(tipo=tipo_filtro)
    if cat_filtro:
        registros = registros.filter(categoria=cat_filtro)

    no_sub = registros.filter(es_subtotal=False)
    totales = {}
    for mes in MESES:
        totales[mes] = _sum(no_sub, mes)
    totales['total'] = _sum(no_sub, 'total')

    context = {
        'registros': registros, 'vigencia': vigencia,
        'tipo_filtro': tipo_filtro, 'cat_filtro': cat_filtro,
        'totales': totales, 'meses': MESES, 'meses_display': MESES_DISPLAY,
        'modulo': 'PAC Ejecutado - Compromisos', 'color': '#2196f3',
    }
    return render(request, 'pac/pac_mensual.html', context)


@login_required
def importar_pac_compromisos(request):
    if request.method == 'POST':
        form = ImportarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = form.cleaned_data['vigencia']
            try:
                count = importar_excel_pac(
                    archivo, vigencia, PACEjecutadoCompromiso, request.user,
                    nombre_hoja='EJECUTADO COMPROMISO'
                )
                CargaArchivo.objects.create(
                    tipo='EJECUTADO_COMPROMISO', archivo=archivo, usuario=request.user,
                    registros_cargados=count,
                    observaciones=f'Vigencia {vigencia}. {count} registros. Hoja: PAC EJECUTADO COMPROMISOS.'
                )
                messages.success(request, f'Se cargaron {count} registros de PAC Ejecutado Compromisos.')
                return redirect('pac_ejecutado_compromisos')
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
    else:
        form = ImportarArchivoForm()
    context = {
        'form': form, 'modulo': 'PAC Ejecutado - Compromisos', 'color': '#2196f3',
        'tipo_importacion': 'EJECUTADO_COMPROMISO',
        'descripcion_formato': 'Se lee la hoja "PAC EJECUTADO COMPROMISOS". Cargar mensualmente al cierre de cifras.',
    }
    return render(request, 'pac/importar.html', context)


# ============================================================
# PAC EJECUTADO PAGOS
# ============================================================
@login_required
def pac_ejecutado_pagos(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    tipo_filtro = request.GET.get('tipo', '')
    cat_filtro = request.GET.get('categoria', '')

    registros = PACEjecutadoPago.objects.filter(vigencia=vigencia)
    if tipo_filtro:
        registros = registros.filter(tipo=tipo_filtro)
    if cat_filtro:
        registros = registros.filter(categoria=cat_filtro)

    no_sub = registros.filter(es_subtotal=False)
    totales = {}
    for mes in MESES:
        totales[mes] = _sum(no_sub, mes)
    totales['total'] = _sum(no_sub, 'total')

    context = {
        'registros': registros, 'vigencia': vigencia,
        'tipo_filtro': tipo_filtro, 'cat_filtro': cat_filtro,
        'totales': totales, 'meses': MESES, 'meses_display': MESES_DISPLAY,
        'modulo': 'PAC Ejecutado - Pagos', 'color': '#9c27b0',
    }
    return render(request, 'pac/pac_mensual.html', context)


@login_required
def importar_pac_pagos(request):
    if request.method == 'POST':
        form = ImportarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            vigencia = form.cleaned_data['vigencia']
            try:
                count = importar_excel_pac(
                    archivo, vigencia, PACEjecutadoPago, request.user,
                    nombre_hoja='EJECUTADO PAGO'
                )
                CargaArchivo.objects.create(
                    tipo='EJECUTADO_PAGO', archivo=archivo, usuario=request.user,
                    registros_cargados=count,
                    observaciones=f'Vigencia {vigencia}. {count} registros. Hoja: PAC EJECUTADO PAGOS.'
                )
                messages.success(request, f'Se cargaron {count} registros de PAC Ejecutado Pagos.')
                return redirect('pac_ejecutado_pagos')
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
    else:
        form = ImportarArchivoForm()
    context = {
        'form': form, 'modulo': 'PAC Ejecutado - Pagos', 'color': '#9c27b0',
        'tipo_importacion': 'EJECUTADO_PAGO',
        'descripcion_formato': 'Se lee la hoja "PAC EJECUTADO PAGOS". Cargar mensualmente al cierre de cifras.',
    }
    return render(request, 'pac/importar.html', context)


# ============================================================
# SEGUIMIENTO PAC
# ============================================================
def _build_seguimiento(vigencia, tipo_pac, modelo_prog, modelo_ejec, label_prog='Programado', label_ejec='Ejecutado'):
    """Construye datos de seguimiento agrupados por categoria con detalle de items."""
    categorias = set()
    for model in [modelo_prog, modelo_ejec]:
        categorias.update(
            model.objects.filter(vigencia=vigencia, tipo=tipo_pac, es_subtotal=False)
            .values_list('categoria', flat=True).distinct()
        )

    datos = []
    for cat in sorted(categorias):
        if not cat:
            continue
        cat_display = dict(AIMInicial.CATEGORIA_CHOICES).get(cat, cat)

        # Fila agregada de categoria
        fila = {'fuente': cat_display, 'es_categoria': True, 'meses': [], 'items': []}
        for mes in MESES:
            prog = _sum(modelo_prog.objects.filter(vigencia=vigencia, tipo=tipo_pac, categoria=cat, es_subtotal=False), mes)
            ejec = _sum(modelo_ejec.objects.filter(vigencia=vigencia, tipo=tipo_pac, categoria=cat, es_subtotal=False), mes)
            pct = (float(ejec) / float(prog) * 100) if prog else 0
            fila['meses'].append({'programado': prog, 'ejecutado': ejec, 'pct': round(pct, 1)})

        prog_total = _sum(modelo_prog.objects.filter(vigencia=vigencia, tipo=tipo_pac, categoria=cat, es_subtotal=False), 'total')
        ejec_total = _sum(modelo_ejec.objects.filter(vigencia=vigencia, tipo=tipo_pac, categoria=cat, es_subtotal=False), 'total')
        pct_total = (float(ejec_total) / float(prog_total) * 100) if prog_total else 0
        fila['prog_total'] = prog_total
        fila['ejec_total'] = ejec_total
        fila['pct_total'] = round(pct_total, 1)

        # Items individuales (rubros hoja) dentro de esta categoria
        rubros = modelo_prog.objects.filter(
            vigencia=vigencia, tipo=tipo_pac, categoria=cat, es_subtotal=False
        ).values_list('codigo_rubro', 'nombre_rubro').distinct()

        for codigo_rubro, nombre_rubro in rubros:
            item = {
                'fuente': nombre_rubro or codigo_rubro,
                'codigo': codigo_rubro,
                'es_categoria': False,
                'meses': [],
            }
            for mes in MESES:
                p = _sum(modelo_prog.objects.filter(
                    vigencia=vigencia, tipo=tipo_pac, categoria=cat,
                    codigo_rubro=codigo_rubro, es_subtotal=False), mes)
                e = _sum(modelo_ejec.objects.filter(
                    vigencia=vigencia, tipo=tipo_pac, categoria=cat,
                    codigo_rubro=codigo_rubro, es_subtotal=False), mes)
                pct_i = (float(e) / float(p) * 100) if p else 0
                item['meses'].append({'programado': p, 'ejecutado': e, 'pct': round(pct_i, 1)})

            pt = _sum(modelo_prog.objects.filter(
                vigencia=vigencia, tipo=tipo_pac, categoria=cat,
                codigo_rubro=codigo_rubro, es_subtotal=False), 'total')
            et = _sum(modelo_ejec.objects.filter(
                vigencia=vigencia, tipo=tipo_pac, categoria=cat,
                codigo_rubro=codigo_rubro, es_subtotal=False), 'total')
            pct_t = (float(et) / float(pt) * 100) if pt else 0
            item['prog_total'] = pt
            item['ejec_total'] = et
            item['pct_total'] = round(pct_t, 1)
            fila['items'].append(item)

        datos.append(fila)

    return datos


def _flat_items(datos):
    """Extrae lista plana de items individuales con su categoria para tabla desagregada."""
    items = []
    for fila in datos:
        for item in fila.get('items', []):
            items.append({**item, 'categoria': fila['fuente']})
    return items


@login_required
def seguimiento_ingresos(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    datos = _build_seguimiento(vigencia, 'INGRESO', PACProgramado, PACEjecutadoPago)
    total_prog = sum(float(f['prog_total']) for f in datos)
    total_ejec = sum(float(f['ejec_total']) for f in datos)
    pct_general = round(total_ejec / total_prog * 100, 1) if total_prog else 0
    context = {
        'datos': datos, 'vigencia': vigencia, 'meses_display': MESES_DISPLAY,
        'items_desagregados': _flat_items(datos),
        'pct_general': pct_general,
        'total_prog': total_prog, 'total_ejec': total_ejec,
        'titulo': 'Seguimiento PAC Ingresos (Recaudo)',
        'color': '#4caf50', 'subtitulo': 'Programado vs Ejecutado (Pagos/Recaudo)',
    }
    return render(request, 'pac/seguimiento.html', context)


@login_required
def seguimiento_gastos(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    datos = _build_seguimiento(vigencia, 'GASTO', PACProgramado, PACEjecutadoPago)
    total_prog = sum(float(f['prog_total']) for f in datos)
    total_ejec = sum(float(f['ejec_total']) for f in datos)
    pct_general = round(total_ejec / total_prog * 100, 1) if total_prog else 0
    context = {
        'datos': datos, 'vigencia': vigencia, 'meses_display': MESES_DISPLAY,
        'items_desagregados': _flat_items(datos),
        'pct_general': pct_general,
        'total_prog': total_prog, 'total_ejec': total_ejec,
        'titulo': 'Seguimiento PAC Gastos (Pagos)',
        'color': '#4caf50', 'subtitulo': 'Programado vs Ejecutado (Pagos)',
    }
    return render(request, 'pac/seguimiento.html', context)


@login_required
def seguimiento_compromisos_vs_pagos(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    datos = _build_seguimiento(vigencia, 'GASTO', PACEjecutadoCompromiso, PACEjecutadoPago)
    context = {
        'datos': datos, 'vigencia': vigencia, 'meses_display': MESES_DISPLAY,
        'titulo': 'Seguimiento Compromisos vs Pagos (Gastos)',
        'color': '#ff5722', 'subtitulo': 'Compromisos vs Pagos',
    }
    return render(request, 'pac/seguimiento.html', context)


# ============================================================
# REPORTES Y ANALISIS
# ============================================================
@login_required
def reportes(request):
    vigencia = int(request.GET.get('vigencia', 2026))

    categorias = set()
    for model in [AIMInicial, PACProgramado, PACEjecutadoCompromiso, PACEjecutadoPago]:
        categorias.update(
            model.objects.filter(vigencia=vigencia, es_subtotal=False)
            .values_list('categoria', flat=True).distinct()
        )

    cat_display_map = dict(AIMInicial.CATEGORIA_CHOICES)
    reporte_fuentes = []
    for cat in sorted(categorias):
        if not cat:
            continue
        fila = {'fuente': cat_display_map.get(cat, cat)}
        base_f = {'vigencia': vigencia, 'categoria': cat, 'es_subtotal': False}

        fila['aim_ingresos'] = _sum(AIMInicial.objects.filter(tipo='INGRESO', **base_f), 'apropiacion_definitiva')
        fila['aim_gastos'] = _sum(AIMInicial.objects.filter(tipo='GASTO', **base_f), 'apropiacion_definitiva')
        fila['prog_ingresos'] = _sum(PACProgramado.objects.filter(tipo='INGRESO', **base_f), 'total')
        fila['prog_gastos'] = _sum(PACProgramado.objects.filter(tipo='GASTO', **base_f), 'total')
        fila['comp_ingresos'] = _sum(PACEjecutadoCompromiso.objects.filter(tipo='INGRESO', **base_f), 'total')
        fila['comp_gastos'] = _sum(PACEjecutadoCompromiso.objects.filter(tipo='GASTO', **base_f), 'total')
        fila['pago_ingresos'] = _sum(PACEjecutadoPago.objects.filter(tipo='INGRESO', **base_f), 'total')
        fila['pago_gastos'] = _sum(PACEjecutadoPago.objects.filter(tipo='GASTO', **base_f), 'total')

        fila['pct_ing'] = round(float(fila['pago_ingresos']) / float(fila['prog_ingresos']) * 100, 1) if fila['prog_ingresos'] else 0
        fila['pct_gas_comp'] = round(float(fila['comp_gastos']) / float(fila['prog_gastos']) * 100, 1) if fila['prog_gastos'] else 0
        fila['pct_gas_pago'] = round(float(fila['pago_gastos']) / float(fila['prog_gastos']) * 100, 1) if fila['prog_gastos'] else 0
        reporte_fuentes.append(fila)

    grafica_fuentes = {
        'labels': [f['fuente'][:30] for f in reporte_fuentes],
        'prog_ingresos': [float(f['prog_ingresos']) for f in reporte_fuentes],
        'prog_gastos': [float(f['prog_gastos']) for f in reporte_fuentes],
        'comp_gastos': [float(f['comp_gastos']) for f in reporte_fuentes],
        'pago_gastos': [float(f['pago_gastos']) for f in reporte_fuentes],
    }

    resumen_mensual = []
    acum_prog_ing = acum_ejec_ing = acum_prog_gas = acum_comp_gas = acum_pago_gas = D0
    base_no_sub = {'vigencia': vigencia, 'es_subtotal': False}
    _excl_ing = ['SALDO_INICIAL']
    _excl_gas = ['RESERVAS', 'CUENTAS_POR_PAGAR']

    for i, mes in enumerate(MESES):
        prog_ing = _sum(PACProgramado.objects.filter(tipo='INGRESO', **base_no_sub).exclude(categoria__in=_excl_ing), mes)
        ejec_ing = _sum(PACEjecutadoPago.objects.filter(tipo='INGRESO', **base_no_sub).exclude(categoria__in=_excl_ing), mes)
        prog_gas = _sum(PACProgramado.objects.filter(tipo='GASTO', **base_no_sub).exclude(categoria__in=_excl_gas), mes)
        comp_gas = _sum(PACEjecutadoCompromiso.objects.filter(tipo='GASTO', **base_no_sub).exclude(categoria__in=_excl_gas), mes)
        pago_gas = _sum(PACEjecutadoPago.objects.filter(tipo='GASTO', **base_no_sub).exclude(categoria__in=_excl_gas), mes)

        acum_prog_ing += prog_ing
        acum_ejec_ing += ejec_ing
        acum_prog_gas += prog_gas
        acum_comp_gas += comp_gas
        acum_pago_gas += pago_gas

        resumen_mensual.append({
            'mes': MESES_DISPLAY[i],
            'prog_ing': prog_ing, 'ejec_ing': ejec_ing,
            'acum_prog_ing': acum_prog_ing, 'acum_ejec_ing': acum_ejec_ing,
            'pct_ing': round(float(ejec_ing) / float(prog_ing) * 100, 1) if prog_ing else 0,
            'prog_gas': prog_gas, 'comp_gas': comp_gas, 'pago_gas': pago_gas,
            'acum_prog_gas': acum_prog_gas, 'acum_comp_gas': acum_comp_gas, 'acum_pago_gas': acum_pago_gas,
            'pct_comp': round(float(comp_gas) / float(prog_gas) * 100, 1) if prog_gas else 0,
            'pct_pago': round(float(pago_gas) / float(prog_gas) * 100, 1) if prog_gas else 0,
        })

    cargas = CargaArchivo.objects.all()[:20]

    context = {
        'vigencia': vigencia, 'reporte_fuentes': reporte_fuentes,
        'grafica_fuentes': json.dumps(grafica_fuentes),
        'resumen_mensual': resumen_mensual,
        'meses_display': json.dumps(MESES_DISPLAY),
        'cargas': cargas,
    }
    return render(request, 'pac/reportes.html', context)


# ============================================================
# EXPORTAR A EXCEL
# ============================================================
@login_required
def exportar_seguimiento_excel(request, tipo):
    vigencia = int(request.GET.get('vigencia', 2026))
    wb = Workbook()
    ws = wb.active
    hf = Font(bold=True, color='FFFFFF', size=11)
    fill_g = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    fill_o = PatternFill(start_color='FF5722', end_color='FF5722', fill_type='solid')
    brd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    if tipo == 'ingresos':
        ws.title = 'Seg. Ingresos'
        fill = fill_g
        datos = _build_seguimiento(vigencia, 'INGRESO', PACProgramado, PACEjecutadoPago)
        titulo = 'SEGUIMIENTO PAC INGRESOS'
    elif tipo == 'gastos':
        ws.title = 'Seg. Gastos'
        fill = fill_g
        datos = _build_seguimiento(vigencia, 'GASTO', PACProgramado, PACEjecutadoPago)
        titulo = 'SEGUIMIENTO PAC GASTOS'
    else:
        ws.title = 'Comp. vs Pagos'
        fill = fill_o
        datos = _build_seguimiento(vigencia, 'GASTO', PACEjecutadoCompromiso, PACEjecutadoPago)
        titulo = 'SEGUIMIENTO COMPROMISOS VS PAGOS'

    ws.append([titulo + f' - Vigencia {vigencia}'])
    headers = ['Categoria']
    for m in MESES_DISPLAY:
        headers.extend([f'{m} Prog.', f'{m} Ejec.', f'{m} %'])
    headers.extend(['Total Prog.', 'Total Ejec.', 'Total %'])
    ws.append(headers)
    for cell in ws[2]:
        cell.font = hf
        cell.fill = fill
        cell.border = brd
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for fila in datos:
        row = [fila['fuente']]
        for md in fila['meses']:
            row.extend([float(md['programado']), float(md['ejecutado']), md['pct']])
        row.extend([float(fila['prog_total']), float(fila['ejec_total']), fila['pct_total']])
        ws.append(row)

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=seguimiento_{tipo}_{vigencia}.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_reporte_fuentes_excel(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte Categorias'
    hf = Font(bold=True, color='FFFFFF', size=11)
    fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    brd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append([f'REPORTE POR CATEGORIAS - Vigencia {vigencia}'])
    headers = ['Categoria', 'AIM Ing.', 'AIM Gas.', 'Prog. Ing.', 'Prog. Gas.',
               'Comp. Rec.', 'Comp. Gas.', 'Pagos Rec.', 'Pagos Gas.',
               '% Ejec. Ing.', '% Ejec. Gas.(C)', '% Ejec. Gas.(P)']
    ws.append(headers)
    for cell in ws[2]:
        cell.font = hf
        cell.fill = fill
        cell.border = brd
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    cat_display_map = dict(AIMInicial.CATEGORIA_CHOICES)
    categorias = set()
    for model in [AIMInicial, PACProgramado, PACEjecutadoCompromiso, PACEjecutadoPago]:
        categorias.update(model.objects.filter(vigencia=vigencia, es_subtotal=False).values_list('categoria', flat=True).distinct())

    for cat in sorted(categorias):
        if not cat:
            continue
        bf = {'vigencia': vigencia, 'categoria': cat, 'es_subtotal': False}
        ai = float(_sum(AIMInicial.objects.filter(tipo='INGRESO', **bf), 'apropiacion_definitiva'))
        ag = float(_sum(AIMInicial.objects.filter(tipo='GASTO', **bf), 'apropiacion_definitiva'))
        pi = float(_sum(PACProgramado.objects.filter(tipo='INGRESO', **bf), 'total'))
        pg = float(_sum(PACProgramado.objects.filter(tipo='GASTO', **bf), 'total'))
        ci = float(_sum(PACEjecutadoCompromiso.objects.filter(tipo='INGRESO', **bf), 'total'))
        cg = float(_sum(PACEjecutadoCompromiso.objects.filter(tipo='GASTO', **bf), 'total'))
        pai = float(_sum(PACEjecutadoPago.objects.filter(tipo='INGRESO', **bf), 'total'))
        pag = float(_sum(PACEjecutadoPago.objects.filter(tipo='GASTO', **bf), 'total'))
        ws.append([
            cat_display_map.get(cat, cat), ai, ag, pi, pg, ci, cg, pai, pag,
            round(pai / pi * 100, 1) if pi else 0,
            round(cg / pg * 100, 1) if pg else 0,
            round(pag / pg * 100, 1) if pg else 0,
        ])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reporte_categorias_{vigencia}.xlsx'
    wb.save(response)
    return response


# ============================================================
# DESCARGAR PLANTILLAS DE EJEMPLO (basadas en los datos reales)
# ============================================================
@login_required
def descargar_plantilla(request, tipo):
    """Genera una plantilla Excel con el mismo formato que los archivos reales de la entidad."""
    wb = Workbook()
    ws = wb.active
    brd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    hf = Font(bold=True, color='FFFFFF', size=11)
    hf_dark = Font(bold=True, size=11)
    section_font = Font(bold=True, size=10)

    if tipo == 'aim_inicial':
        ws.title = 'AIM INICIAL'
        fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
        titulo = 'PROGRAMACION PLAN ANUAL DE CAJA 2026 - AIM INICIAL'
        filename = 'plantilla_aim_inicial.xlsx'
    elif tipo == 'programado':
        ws.title = 'PROG PAC INGRESOS-GASTOS 2026'
        fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
        titulo = 'PROGRAMACION PLAN ANUAL DE CAJA 2026'
        filename = 'plantilla_pac_programado.xlsx'
    elif tipo == 'compromisos':
        ws.title = 'PAC EJECUTADO COMPROMISOS'
        fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        titulo = 'EJECUCION (COMPROMISOS) PLAN ANUAL DE CAJA 2026'
        filename = 'plantilla_pac_ejecutado_compromisos.xlsx'
    elif tipo == 'pagos':
        ws.title = 'PAC EJECUTADO PAGOS'
        fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
        titulo = 'EJECUCION (PAGOS) PLAN ANUAL DE CAJA 2026'
        filename = 'plantilla_pac_ejecutado_pagos.xlsx'
    else:
        return HttpResponse('Tipo no valido', status=400)

    # Fila 1: Titulo
    ws.cell(row=1, column=3, value=titulo).font = Font(bold=True, size=14)
    ws.merge_cells('C1:I1')
    # Fila 2: Entidad
    ws.cell(row=2, column=3, value='ENTIDAD EJEMPLO').font = Font(bold=True, size=12)
    ws.merge_cells('C2:I2')
    # Fila 3: vacia

    # Fila 4: Encabezados
    headers = {
        'B': 'CODIGO', 'C': 'INGRESOS',
        'D': 'Apro Inicial', 'E': 'Adiciones', 'F': 'Reduccion',
        'G': 'Creditos', 'H': 'Contracred', 'I': 'Apro Definitiva',
        'J': 'Enero', 'K': 'Febrero', 'L': 'Marzo', 'M': 'Abril',
        'N': 'Mayo', 'O': 'Junio', 'P': 'Julio', 'Q': 'Agosto',
        'R': 'Septiembre', 'S': 'Octubre', 'T': 'Noviembre', 'U': 'Diciembre',
        'V': 'Total'
    }
    for col_letter, header in headers.items():
        cell = ws[f'{col_letter}4']
        cell.value = header
        cell.font = hf
        cell.fill = fill
        cell.border = brd
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Datos de ejemplo basados en la estructura real
    ejemplo_data = [
        # Fila, CodB, DescC, AproIni, Adic, Reduc, Cred, Contrac, AproDef, Ene-Dic..., Total
        (5, '1', 'Saldo Inicial', 90641722287.64, 0, 0, 0, 0, 90641722287.64,
         0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
        (6, '1.1', '  Caja', 0, 0, 0, 0, 0, 0,
         0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
        (7, '1.2', '  Bancos', 90641722287.64, 0, 0, 0, 0, 90641722287.64,
         0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
        (9, '2', 'Ingresos Corrientes', 101909465820.87, 0, 0, 0, 0, 101909465820.87,
         0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
        (10, '', '  Tributarios', 101909465820.87, 0, 0, 0, 0, 101909465820.87,
         0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
        (13, '1003 - 1.1.02.06.006.06 01 - 03', '    Funcionamiento', 11768629792.90, 0, 0, 0, 0, 11768629792.90,
         980719149.41, 980719149.41, 980719149.41, 980719149.41, 980719149.41, 980719149.41,
         980719149.41, 980719149.41, 980719149.41, 980719149.41, 980719149.41, 980719149.41, 11768629792.90),
        (14, '1003 - 1.2.08.06.002.01 - XX', '    Transferencias para Inversion', 90140836027.97, 0, 0, 0, 0, 90140836027.97,
         7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66,
         7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 90140836027.97),
        (28, 'A', 'Total Ingresos (1+2+3)', 192551188108.51, 0, 0, 0, 0, 192551188108.51,
         8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07,
         8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 101909465820.87),
    ]

    # Seccion GASTOS - cambiar encabezado
    gastos_data = [
        (29, '', 'GASTOS', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
        (30, '1003 - 2.1', 'Funcionamiento', 11768629792.90, 0, 0, 0, 0, 11768629792.90,
         980719149.41, 980719149.41, 980719149.41, 980719149.41, 980719149.41, 980719149.41,
         980719149.41, 980719149.41, 980719149.41, 980719149.41, 980719149.41, 980719149.41, 11768629792.90),
        (31, '1003 - 2.1.1.01.01.001.01 - 20', '  Sueldo basico', 5543330440.82, 0, 0, 0, 0, 5543330440.82,
         461944203.40, 461944203.40, 461944203.40, 461944203.40, 461944203.40, 461944203.40,
         461944203.40, 461944203.40, 461944203.40, 461944203.40, 461944203.40, 461944203.40, 5543330440.82),
        (32, '1003 - 2.1.1.01.01.004.01 - 20', '  Prima de servicio', 380000000, 0, 0, 0, 0, 380000000,
         0, 0, 0, 0, 0, 190000000, 0, 0, 0, 0, 0, 190000000, 380000000),
        (74, '1003 - 2.3', 'Inversion', 90140836027.97, 0, 0, 0, 0, 90140836027.97,
         7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66,
         7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 7511736335.66, 90140836027.97),
        (75, '', 'SECTOR MINAS Y ENERGIA', 8248505682.57, 0, 0, 0, 0, 8248505682.57,
         687375473.55, 687375473.55, 687375473.55, 687375473.55, 687375473.55, 687375473.55,
         687375473.55, 687375473.55, 687375473.55, 687375473.55, 687375473.55, 687375473.55, 8248505682.57),
        (76, '1003 - 2.3.21.2024005500143 - 23', '  BPIN 2024005500143 Generacion Fotovoltaica', 2380952380.95, 0, 0, 0, 0, 2380952380.95,
         198412698.41, 198412698.41, 198412698.41, 198412698.41, 198412698.41, 198412698.41,
         198412698.41, 198412698.41, 198412698.41, 198412698.41, 198412698.41, 198412698.41, 2380952380.95),
        (260, 'B', 'Total Gastos', 101909465820.87, 0, 0, 0, 0, 101909465820.87,
         8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07,
         8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 8492455485.07, 101909465820.87),
        (261, '', 'Saldo Disponible (A - B)', 90641722287.64, 0, 0, 0, 0, 90641722287.64,
         0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
    ]

    all_data = ejemplo_data + gastos_data

    # Estilos
    sub_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    gastos_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')

    for data_row in all_data:
        fila_num = data_row[0]
        ws.cell(row=fila_num, column=2, value=data_row[1])  # Codigo
        ws.cell(row=fila_num, column=3, value=data_row[2])  # Descripcion
        for col_idx in range(3, 22):
            val = data_row[col_idx]
            ws.cell(row=fila_num, column=col_idx + 1, value=val)
            ws.cell(row=fila_num, column=col_idx + 1).number_format = '#,##0.00'

        # Estilo para filas especiales
        desc = data_row[2].strip().upper()
        if desc == 'GASTOS':
            for c in range(2, 23):
                ws.cell(row=fila_num, column=c).fill = gastos_fill
                ws.cell(row=fila_num, column=c).font = section_font
        elif data_row[1] in ('A', 'B') or 'TOTAL' in desc:
            for c in range(2, 23):
                ws.cell(row=fila_num, column=c).font = Font(bold=True)
                ws.cell(row=fila_num, column=c).fill = sub_fill

    # Ajustar anchos
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 45
    for col_letter in 'DEFGHIJKLMNOPQRSTUV':
        ws.column_dimensions[col_letter].width = 16

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


# ============================================================
# ELIMINAR DATOS
# ============================================================
@login_required
def eliminar_datos(request, tipo):
    vigencia = int(request.GET.get('vigencia', 2026))
    if request.method == 'POST':
        modelos = {
            'aim_inicial': (AIMInicial, 'AIM Inicial'),
            'programado': (PACProgramado, 'PAC Programado'),
            'compromisos': (PACEjecutadoCompromiso, 'PAC Ejecutado Compromisos'),
            'pagos': (PACEjecutadoPago, 'PAC Ejecutado Pagos'),
        }
        if tipo in modelos:
            modelo, nombre = modelos[tipo]
            count = modelo.objects.filter(vigencia=vigencia).count()
            modelo.objects.filter(vigencia=vigencia).delete()
            messages.success(request, f'Se eliminaron {count} registros de {nombre} (Vigencia {vigencia}).')
    return redirect(request.META.get('HTTP_REFERER', '/'))


# ============================================================
# FUENTES DE FINANCIACION (CRUD)
# ============================================================
@login_required
def fuentes_financiacion(request):
    vigencia = int(request.GET.get('vigencia', 2026))
    fuentes = FuenteFinanciacion.objects.filter(vigencia=vigencia)
    fuentes_data = []
    for fuente in fuentes:
        fuentes_data.append({
            'obj': fuente,
            'programado_ing': fuente.get_total_programado_ingresos(),
            'programado_gas': fuente.get_total_programado_gastos(),
            'compromisos': fuente.get_total_compromisos(),
            'pagos': fuente.get_total_pagos_gastos(),
            'recaudo': fuente.get_total_recaudo(),
            'saldo': fuente.get_saldo_disponible(),
            'pct_ejecucion': fuente.get_porcentaje_ejecucion(),
            'pct_pagos': fuente.get_porcentaje_pagos(),
        })
    context = {'fuentes_data': fuentes_data, 'vigencia': vigencia}
    return render(request, 'pac/fuentes.html', context)


@login_required
def fuente_crear(request):
    if request.method == 'POST':
        form = FuenteFinanciacionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fuente de financiacion creada correctamente.')
            return redirect('fuentes_financiacion')
    else:
        form = FuenteFinanciacionForm()
    return render(request, 'pac/fuente_form.html', {'form': form, 'titulo': 'Crear Fuente de Financiacion'})


@login_required
def fuente_editar(request, pk):
    fuente = get_object_or_404(FuenteFinanciacion, pk=pk)
    if request.method == 'POST':
        form = FuenteFinanciacionForm(request.POST, instance=fuente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fuente de financiacion actualizada correctamente.')
            return redirect('fuentes_financiacion')
    else:
        form = FuenteFinanciacionForm(instance=fuente)
    return render(request, 'pac/fuente_form.html', {'form': form, 'titulo': 'Editar Fuente de Financiacion'})


@login_required
def fuente_eliminar(request, pk):
    fuente = get_object_or_404(FuenteFinanciacion, pk=pk)
    if request.method == 'POST':
        fuente.delete()
        messages.success(request, 'Fuente de financiacion eliminada correctamente.')
    return redirect('fuentes_financiacion')


@login_required
def fuente_detalle(request, pk):
    fuente = get_object_or_404(FuenteFinanciacion, pk=pk)
    datos_mensuales = []
    for i, mes in enumerate(MESES):
        prog_ing = _sum(PACProgramado.objects.filter(vigencia=fuente.vigencia, tipo='INGRESO', fuente_financiacion=fuente.nombre, es_subtotal=False), mes)
        prog_gas = _sum(PACProgramado.objects.filter(vigencia=fuente.vigencia, tipo='GASTO', fuente_financiacion=fuente.nombre, es_subtotal=False), mes)
        comp = _sum(PACEjecutadoCompromiso.objects.filter(vigencia=fuente.vigencia, tipo='GASTO', fuente_financiacion=fuente.nombre, es_subtotal=False), mes)
        pago = _sum(PACEjecutadoPago.objects.filter(vigencia=fuente.vigencia, tipo='GASTO', fuente_financiacion=fuente.nombre, es_subtotal=False), mes)
        recaudo = _sum(PACEjecutadoPago.objects.filter(vigencia=fuente.vigencia, tipo='INGRESO', fuente_financiacion=fuente.nombre, es_subtotal=False), mes)
        datos_mensuales.append({
            'mes': MESES_DISPLAY[i],
            'prog_ing': prog_ing, 'prog_gas': prog_gas,
            'compromisos': comp, 'pagos': pago, 'recaudo': recaudo,
            'pct_comp': round(float(comp) / float(prog_gas) * 100, 1) if prog_gas else 0,
            'pct_pago': round(float(pago) / float(prog_gas) * 100, 1) if prog_gas else 0,
        })

    rubros_ingreso = PACProgramado.objects.filter(vigencia=fuente.vigencia, tipo='INGRESO', fuente_financiacion=fuente.nombre, es_subtotal=False).values('codigo_rubro', 'nombre_rubro', 'total')
    rubros_gasto = PACProgramado.objects.filter(vigencia=fuente.vigencia, tipo='GASTO', fuente_financiacion=fuente.nombre, es_subtotal=False).values('codigo_rubro', 'nombre_rubro', 'total')

    context = {
        'fuente': fuente, 'datos_mensuales': datos_mensuales,
        'meses_display': json.dumps(MESES_DISPLAY),
        'rubros_ingreso': rubros_ingreso, 'rubros_gasto': rubros_gasto,
        'total_programado_ing': fuente.get_total_programado_ingresos(),
        'total_programado_gas': fuente.get_total_programado_gastos(),
        'total_compromisos': fuente.get_total_compromisos(),
        'total_pagos': fuente.get_total_pagos_gastos(),
        'total_recaudo': fuente.get_total_recaudo(),
        'saldo': fuente.get_saldo_disponible(),
        'pct_ejecucion': fuente.get_porcentaje_ejecucion(),
    }
    return render(request, 'pac/fuente_detalle.html', context)
